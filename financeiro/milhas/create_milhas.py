#!/usr/bin/env python3
"""
Milhas v3 - Geracao limpa
  Python  -> aba dados (Tabela Excel) + aba estoque
  VBA     -> aba anual + named ranges + botoes (execute Setup)

NOTA: Named ranges NAO sao definidos aqui — o Excel Mac rejeita o arquivo
      quando o openpyxl gera XML de named ranges com formulas INDEX/COUNTA.
      O VBA os define em runtime via DefinirNamedRanges() chamado por Setup.
"""
import json
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime, date

AZ_ESC  = "1B3A6B"; AZ_MED  = "2C3E50"; AZ_MED2 = "1A5276"
VERDE   = "1A7A4A"; VD_ESC  = "2C5F2E"
BRANCO  = "FFFFFF"; AZ_CL   = "EBF5FB"; AMAR    = "FFFDE7"
CI_MED  = "BDC3C7"; AZ_ESC2 = "1B3A6B"

FMT_DATA = 'DD/MM/YYYY'
FMT_DEC  = '#,##0.00;[Red]-#,##0.00'
FMT_INT  = '#,##0'
FMT_MILL = '#,##0.00'
ROW_HT   = 16

_b = Side(style='thin', color=CI_MED)
BORDA = Border(left=_b, right=_b, top=_b, bottom=_b)

def cel_h(ws, row, col, val, bg=AZ_ESC, fg=BRANCO, sz=11,
          bold=True, halign='center', cs=1, rs=1):
    c = ws.cell(row=row, column=col, value=val)
    c.font      = Font(name='Arial', size=sz, bold=bold, color=fg)
    c.fill      = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=True)
    c.border    = BORDA
    if cs > 1 or rs > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row+rs-1, end_column=col+cs-1)
    return c

def cel_v(ws, row, col, val=None, fmt=None, halign='right',
          fg='000000', bg=None, bold=False, sz=11):
    c = ws.cell(row=row, column=col)
    if val is not None:
        c.value = val
    c.font      = Font(name='Arial', size=sz, bold=bold, color=fg)
    c.alignment = Alignment(horizontal=halign, vertical='center')
    c.border    = BORDA
    if fmt:
        c.number_format = fmt
    if bg:
        c.fill = PatternFill('solid', start_color=bg)
    return c

# ── Carga de dados ────────────────────────────────────────────────────────────
SRC = '/mnt/user-data/uploads/milhas_-_na_o_apagar.xlsm'
print(f"Carregando {SRC}...")
wb_src   = openpyxl.load_workbook(SRC, keep_vba=False, data_only=True)
src_rows = [r for r in wb_src['dados'].iter_rows(min_row=2, values_only=True)
            if any(x is not None for x in r)]
N = len(src_rows)
print(f"  {N} linhas carregadas.")

def clean_tipo(v):
    if v is None: return None
    s = str(v).strip().upper()
    return 'SAIDA' if 'SA' in s else ('ENTRADA' if 'ENT' in s else v)

# VALOR BRUTO deve ser sempre positivo (ver CLAUDE.md desta pasta).
# Falha cedo em vez de deixar a validacao VBA (ValidarDados) pegar depois.
_val_brutos_invalidos = [
    (i + 2, row[2]) for i, row in enumerate(src_rows)
    if row[2] is not None and row[2] <= 0
]
if _val_brutos_invalidos:
    linhas = ', '.join(f'linha {r} (valor={v})' for r, v in _val_brutos_invalidos)
    raise ValueError(f"VALOR BRUTO deve ser sempre positivo. Corrija na origem: {linhas}")

# ── Novo workbook ─────────────────────────────────────────────────────────────
wb       = Workbook()
ws_dados = wb.active;         ws_dados.title = 'dados'
ws_estq  = wb.create_sheet('estoque')
ws_anual = wb.create_sheet('anual')
wb.create_sheet('tabelas')
wb.create_sheet('hotmilhas')

# ════════════════════════════════════════════════════════════════════════════════
# ABA: dados
# ════════════════════════════════════════════════════════════════════════════════
print("Construindo aba dados...")

HDRS = ['DIA','PAGAMENTO','VALOR BRUTO','MOV. CAIXA','TIPO',
        'PRODUTO','SUBGRUPO','QUANT. PONTOS/MILHAS','VALOR MEDIO MILHEIRO',
        'OBSERVACOES','FAMILIAR','CARTAO','TOTAL PARCELAS','PARCELA ATUAL']
for i, h in enumerate(HDRS, 1):
    cel_h(ws_dados, 1, i, h, sz=12)
ws_dados.row_dimensions[1].height = ROW_HT

for row in src_rows:
    ws_dados.append([
        row[0],              # A  DIA
        row[1],              # B  PAGAMENTO
        row[2],              # C  VALOR BRUTO
        None,                # D  MOV. CAIXA  (formula)
        clean_tipo(row[4]),  # E  TIPO
        row[5],              # F  PRODUTO
        row[6],              # G  SUBGRUPO
        row[7],              # H  QUANT. PONTOS/MILHAS
        None,                # I  VALOR MEDIO MILHEIRO (formula)
        row[9],              # J  OBSERVACOES
        row[10],             # K  FAMILIAR
        row[11],             # L  CARTAO
        row[12],             # M  TOTAL PARCELAS
        row[13],             # N  PARCELA ATUAL
    ])

_fnt12  = Font(name='Arial', size=12)
_aln_c  = Alignment(horizontal='center', vertical='center')
for r in range(2, N + 2):
    for c in range(1, 15):
        ws_dados.cell(r, c).font      = _fnt12
        ws_dados.cell(r, c).alignment = _aln_c
    ws_dados.cell(r, 4).value         = f'=IF(E{r}="SAIDA",-C{r},C{r})'
    ws_dados.cell(r, 4).number_format = FMT_DEC
    # VALOR MEDIO MILHEIRO — simplificado (sempre positivo, sem IF por direcao)
    ws_dados.cell(r, 9).value = (
        f'=IFERROR(IF(AND(H{r}<>"",M{r}<>"",M{r}<>0),'
        f'C{r}*M{r}*1000/H{r},""),"")')
    ws_dados.cell(r, 9).number_format = FMT_MILL
    ws_dados.cell(r, 1).number_format = FMT_DATA
    ws_dados.cell(r, 2).number_format = FMT_DATA
    ws_dados.cell(r, 3).number_format = FMT_DEC
    ws_dados.row_dimensions[r].height = ROW_HT

for col, w in {c: 12 for c in range(1, 15)}.items():
    ws_dados.column_dimensions[get_column_letter(col)].width = w

# Tabela Excel (somente na aba dados)
tbl = Table(displayName="Tabela1", ref=f"A1:N{N+1}")
tbl.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium2", showRowStripes=True,
    showFirstColumn=False, showLastColumn=False)
ws_dados.add_table(tbl)

# Data validation coluna E
dv = DataValidation(type="list", formula1='"ENTRADA,SAIDA"',
                    allow_blank=True, showDropDown=False,
                    showErrorMessage=True,
                    error='Use: ENTRADA ou SAIDA', errorTitle='Tipo invalido')
ws_dados.add_data_validation(dv)
dv.add(f'E2:E{N+1}')

ws_dados.freeze_panes = 'A2'

# ════════════════════════════════════════════════════════════════════════════════
# ABA: estoque  (coordenadas exatas conforme especificacao)
# ════════════════════════════════════════════════════════════════════════════════
print("Construindo aba estoque...")

PRGS = ['azul','latam','smiles','british','copa','iberia',
        'qatar','tap','c6','esfera','livelo']

# Nomes reais de familiares ficam fora do Git — ver titulares_config.json
# (listado no .gitignore) e financeiro/milhas/CLAUDE.md.
_TITS_CONFIG = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'titulares_config.json')
with open(_TITS_CONFIG, encoding='utf-8') as f:
    TITS = json.load(f)['titulares']

# Row 1: cabecalhos
cel_h(ws_estq, 1, 1, 'TITULAR', bg=AZ_MED)
for ci, p in enumerate(PRGS, 2):
    cel_h(ws_estq, 1, ci, p.upper(), bg=AZ_ESC)
cel_h(ws_estq, 1, 13, 'TOTAL', bg=AZ_MED)
ws_estq.row_dimensions[1].height = ROW_HT

# Rows 2-10: titulares (texto livre)
for ri, t in enumerate(TITS, 2):
    cel_v(ws_estq, ri, 1, t.upper(), halign='left', bold=True,
          fg=AZ_ESC, bg='F8F9FA')
    for ci in range(2, 13):
        c = ws_estq.cell(ri, ci)
        c.number_format = FMT_INT
        c.alignment     = Alignment(horizontal='right', vertical='center')
        c.border        = BORDA
    # Total por titular
    c = ws_estq.cell(ri, 13)
    c.value         = f'=SUM(B{ri}:L{ri})'
    c.number_format = FMT_INT
    c.font          = Font(name='Arial', size=11, bold=True, color=AZ_ESC)
    c.fill          = PatternFill('solid', start_color=AZ_CL)
    c.alignment     = Alignment(horizontal='right', vertical='center')
    c.border        = BORDA
    ws_estq.row_dimensions[ri].height = ROW_HT

# Row 11: TOTAL DE MILHAS  =SUM(B2:B10) etc.
cel_h(ws_estq, 11, 1, 'TOTAL DE MILHAS', bg=AZ_ESC)
for ci in range(2, 13):
    L = get_column_letter(ci)
    c = ws_estq.cell(11, ci)
    c.value         = f'=SUM({L}2:{L}10)'
    c.number_format = FMT_INT
    c.font          = Font(name='Arial', size=11, bold=True, color=BRANCO)
    c.fill          = PatternFill('solid', start_color=AZ_ESC)
    c.alignment     = Alignment(horizontal='right', vertical='center')
    c.border        = BORDA
c = ws_estq.cell(11, 13)
c.value         = '=SUM(B11:L11)'
c.number_format = FMT_INT
c.font          = Font(name='Arial', size=11, bold=True, color=BRANCO)
c.fill          = PatternFill('solid', start_color=AZ_MED)
c.alignment     = Alignment(horizontal='right', vertical='center')
c.border        = BORDA
ws_estq.row_dimensions[11].height = ROW_HT

# Row 12: COTACAO (texto livre — usuario preenche)
cel_h(ws_estq, 12, 1, 'COTACAO', bg=VD_ESC)
for ci in range(2, 13):
    c = ws_estq.cell(12, ci)
    c.number_format = FMT_INT
    c.font          = Font(name='Arial', size=11, bold=True, color=AZ_ESC)
    c.fill          = PatternFill('solid', start_color=AMAR)
    c.alignment     = Alignment(horizontal='right', vertical='center')
    c.border        = BORDA
ws_estq.row_dimensions[12].height = ROW_HT

# Row 13: VALOR DO ESTOQUE  =(B11*B12)/1000
cel_h(ws_estq, 13, 1, 'VALOR DO ESTOQUE', bg=AZ_MED2)
for ci in range(2, 13):
    L = get_column_letter(ci)
    c = ws_estq.cell(13, ci)
    c.value         = f'=({L}11*{L}12)/1000'
    c.number_format = FMT_INT
    c.font          = Font(name='Arial', size=12, bold=True, color=BRANCO)
    c.fill          = PatternFill('solid', start_color=AZ_MED2)
    c.alignment     = Alignment(horizontal='right', vertical='center')
    c.border        = BORDA
# M13: TOTAL GERAL ESTOQUE
c = ws_estq.cell(13, 13)
c.value         = '=SUM(B13:L13)'
c.number_format = FMT_INT
c.font          = Font(name='Arial', size=12, bold=True, color=BRANCO)
c.fill          = PatternFill('solid', start_color=VERDE)
c.alignment     = Alignment(horizontal='center', vertical='center')
c.border        = BORDA
ws_estq.row_dimensions[13].height = ROW_HT

ws_estq.column_dimensions['A'].width = 22
for ci in range(2, 13):
    ws_estq.column_dimensions[get_column_letter(ci)].width = 13
ws_estq.column_dimensions['M'].width = 18

# Arial 12 centralizado em toda a aba estoque (preserva cor existente)
for _rw in ws_estq.iter_rows():
    for _cl in _rw:
        _orig = _cl.font
        _bold = _orig.bold if _orig else False
        _clr  = (_orig.color.rgb if _orig and _orig.color and _orig.color.type == 'rgb' else None)
        _cl.font = Font(name='Arial', size=12, bold=_bold,
                        color=_clr if _clr and _clr != '00000000' else '000000')
        _cl.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=_cl.alignment.wrap_text if _cl.alignment else False)
ws_estq.freeze_panes = 'B2'

# ════════════════════════════════════════════════════════════════════════════════
# ABA: anual — INTENCIONALMENTE VAZIA
# O VBA (macro Setup) constroi toda a estrutura + define named ranges
# ════════════════════════════════════════════════════════════════════════════════
ws_anual.cell(1, 1).value = 'Execute a macro "Setup" no VBEditor para construir esta aba.'
ws_anual.cell(1, 1).font  = Font(name='Arial', size=12, color='C0392B', bold=True)
ws_anual.column_dimensions['A'].width = 65

# ── Salvar (SEM named ranges — VBA os define em runtime) ─────────────────────
OUT = '/home/claude/milhas_novo.xlsx'
wb.save(OUT)
print(f"\nSalvo: {OUT}")
print("\nPROXIMOS PASSOS:")
print("  1. Abrir milhas_novo.xlsx  ->  Salvar Como .xlsm")
print("  2. VBEditor  ->  File > Import File  ->  milhas_vba.bas")
print("  3. No VBEditor executar:  Setup")
print("  4. Salvar .xlsm")
